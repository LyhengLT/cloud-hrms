import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Employee, Payslip
from ..schemas import PayslipCreate, PayslipOut
from ..security import get_current_user, require_hr

router = APIRouter(prefix="/api/payroll", tags=["payroll"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _can_view(user: Employee, slip: Payslip) -> bool:
    return user.role.value in ("ADMIN", "HR") or slip.employee_id == user.id


def _to_out(p: Payslip) -> PayslipOut:
    out = PayslipOut.model_validate(p)
    out.employee_name = p.employee.full_name if p.employee else None
    return out


@router.post("", response_model=PayslipOut, status_code=201)
def generate_payslip(
    payload: PayslipCreate,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_hr),
):
    emp = db.query(Employee).get(payload.employee_id)
    if not emp:
        raise HTTPException(404, "Employee not found")

    existing = (
        db.query(Payslip)
        .filter(Payslip.employee_id == emp.id, Payslip.period == payload.period)
        .first()
    )
    if existing:
        raise HTTPException(400, f"Payslip for {payload.period} already exists")

    net = emp.base_salary + payload.allowances - payload.deductions
    slip = Payslip(
        employee_id=emp.id,
        period=payload.period,
        base_salary=emp.base_salary,
        allowances=payload.allowances,
        deductions=payload.deductions,
        net_pay=net,
    )
    db.add(slip)
    db.commit()
    db.refresh(slip)
    return _to_out(slip)


@router.get("/me", response_model=list[PayslipOut])
def my_payslips(db: Session = Depends(get_db), user: Employee = Depends(get_current_user)):
    rows = (
        db.query(Payslip)
        .filter(Payslip.employee_id == user.id)
        .order_by(Payslip.period.desc())
        .all()
    )
    return [_to_out(p) for p in rows]


@router.get("", response_model=list[PayslipOut])
def all_payslips(
    period: str | None = None,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_hr),
):
    query = db.query(Payslip)
    if period:
        query = query.filter(Payslip.period == period)
    rows = query.order_by(Payslip.period.desc()).all()
    return [_to_out(p) for p in rows]


# ---------------------- Exports ----------------------

@router.get("/{slip_id}/pdf")
def payslip_pdf(
    slip_id: int,
    db: Session = Depends(get_db),
    user: Employee = Depends(get_current_user),
):
    """Render a single payslip as a formatted PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    slip = db.query(Payslip).get(slip_id)
    if not slip:
        raise HTTPException(404, "Payslip not found")
    if not _can_view(user, slip):
        raise HTTPException(403, "Insufficient permissions")

    emp = slip.employee
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    x = 25 * mm
    y = h - 30 * mm

    # Header band
    c.setFillColorRGB(0.31, 0.27, 0.90)
    c.rect(0, h - 22 * mm, w, 22 * mm, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x, h - 15 * mm, "Cloud HRMS — Payslip")

    c.setFillColorRGB(0.1, 0.15, 0.24)
    c.setFont("Helvetica", 11)
    y -= 6 * mm
    c.drawString(x, y, f"Employee: {emp.full_name}")
    c.drawString(x + 95 * mm, y, f"Employee ID: #{emp.id}")
    y -= 7 * mm
    c.drawString(x, y, f"Position: {emp.position or '-'}")
    c.drawString(x + 95 * mm, y, f"Pay period: {slip.period}")
    y -= 7 * mm
    c.drawString(x, y, f"Department: {emp.department.name if emp.department else '-'}")
    c.drawString(x + 95 * mm, y, f"Issued: {slip.generated_at:%Y-%m-%d}")

    # Earnings table
    y -= 14 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Earnings & Deductions")
    c.setStrokeColorRGB(0.85, 0.87, 0.92)
    y -= 3 * mm
    c.line(x, y, w - x, y)

    def row(label, value, bold=False):
        nonlocal y
        y -= 8 * mm
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 11)
        c.drawString(x, y, label)
        c.drawRightString(w - x, y, f"${value:,.2f}")

    row("Base salary", slip.base_salary)
    row("Allowances", slip.allowances)
    row("Deductions", -slip.deductions)
    y -= 4 * mm
    c.line(x, y, w - x, y)
    c.setFillColorRGB(0.31, 0.27, 0.90)
    row("NET PAY", slip.net_pay, bold=True)

    c.setFillColorRGB(0.4, 0.45, 0.5)
    c.setFont("Helvetica-Oblique", 9)
    c.drawString(x, 20 * mm, "This payslip was generated electronically by Cloud HRMS.")
    c.showPage()
    c.save()
    buf.seek(0)
    return Response(
        buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="payslip-{slip.period}-{emp.full_name}.pdf"'},
    )


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    user: Employee = Depends(get_current_user),
):
    """Export payslips to a styled Excel workbook (own rows for employees, all for HR/Admin)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    query = db.query(Payslip)
    if user.role.value == "EMPLOYEE":
        query = query.filter(Payslip.employee_id == user.id)
    rows = query.order_by(Payslip.period.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    headers = ["ID", "Employee", "Period", "Base Salary", "Allowances", "Deductions", "Net Pay", "Issued"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="4F46E5")
    for col, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for p in rows:
        ws.append([
            p.id,
            p.employee.full_name if p.employee else "",
            p.period,
            p.base_salary,
            p.allowances,
            p.deductions,
            p.net_pay,
            p.generated_at.strftime("%Y-%m-%d"),
        ])

    for col in ("D", "E", "F", "G"):
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0.00'

    widths = [6, 24, 10, 14, 13, 13, 13, 12]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"payroll-{datetime.utcnow():%Y%m%d}.xlsx"
    return Response(
        buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
